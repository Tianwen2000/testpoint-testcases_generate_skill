from __future__ import annotations

import re
from pathlib import Path
from typing import Callable, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

import generate_v180_testpoint_xmind as testpoint_source


ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = ROOT / "outputs" / "v1.8.0_testcases"
OUTPUT_FILE = OUTPUT_DIR / "音创AI_MelodAI_V1.8.0_首页_创作页_我的_测试用例.xlsx"

PRODUCT = "音创AI/Melod AI"
VERSION = "V1.8.0"
PROJECT = "音创AI/Melod AI V1.8.0"


def walk_leaves(item: dict, path_parts: list[str], out: list[dict]) -> None:
    next_path = [*path_parts, item["title"]]
    children = item.get("children") or []
    if not children:
        out.append({"title": item["title"], "path": next_path})
        return
    for child in children:
        walk_leaves(child, next_path, out)


def section_of(leaf: dict) -> str:
    return leaf["path"][2] if len(leaf["path"]) > 2 else leaf["path"][1]


def module_of(leaf: dict) -> str:
    return leaf["path"][1]


def branch_of(leaf: dict) -> str:
    branch = leaf["path"][2:-1]
    return " -> ".join(branch) if branch else section_of(leaf)


def path_text(leaf: dict) -> str:
    return " > ".join(leaf["path"])


def has(pattern: str, text: str) -> bool:
    return re.search(pattern, text) is not None


def is_pending_title(title: str) -> bool:
    return has(r"待确认|需确认|未定义|冲突|设计图未定义|规则待确认|处理待确认|兜底待确认|范围待确认", title)


def node_type(title: str) -> str:
    if is_pending_title(title):
        return "待确认"
    if has(r"失败|拒绝|不足|异常|弱网|重复|取消|超时|不可用|杀进程|错误", title):
        return "异常"
    if has(r"成功|回流|返回|恢复|跳转|刷新|清空|保留|退回|回显|结果", title):
        return "结果"
    if has(r"状态|排队|生成中|下载中|处理中|置灰|禁用|收起|展开", title):
        return "状态"
    if has(r"展示|显示|隐藏|默认|文案|入口|页面|列表|横幅|弹窗|气泡|Toast", title):
        return "展示"
    if has(r"点击|上传|创建|删除|下载|分享|切换|选择|输入|勾选|关闭|打开|拉起|跳转", title):
        return "交互"
    if has(r"会员|非会员|钻石|扣|消耗|权限|限制|字数|范围|配置|同步|不重复|二次确认", title):
        return "规则"
    return "规则"


def scenario_tag(title: str) -> str:
    if has(r"失败", title):
        return "失败"
    if has(r"不足|无钻石|余额不足", title):
        return "资源不足"
    if has(r"会员", title) and has(r"非会员", title):
        return "会员与非会员分流"
    if has(r"会员", title):
        return "会员路径"
    if has(r"非会员", title):
        return "非会员路径"
    if has(r"权限", title):
        return "权限"
    if has(r"边界|2 分钟|80|200|4500|0-10|50|20 个|1.5 秒|3 秒", title):
        return "边界"
    if has(r"成功", title):
        return "成功"
    if has(r"默认", title):
        return "默认状态"
    if has(r"回流|返回|跳转|恢复", title):
        return "回流"
    if has(r"创建|上传|下载|分轨|创作|添加|删除|分享|切换|点击", title):
        return "交互"
    return "规则"


def unique(items: Iterable[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for item in items:
        if item not in seen:
            seen.add(item)
            result.append(item)
    return result


def preconditions(module: str, title: str) -> str:
    items = ["测试环境已安装 V1.8.0 版本应用"]
    if has(r"非会员", title):
        items.append("已登录非会员账号")
    elif has(r"会员", title):
        items.append("已登录会员账号")
    else:
        items.append("已登录普通测试账号")
    if has(r"钻石充足|余额充足|有钻石|足够", title):
        items.append("账号钻石余额满足本场景消耗")
    if has(r"钻石不足|无钻石|余额不足", title):
        items.append("账号钻石余额不足或为 0")
    if has(r"上传|音频|分轨", title):
        items.append("已准备可用于测试的音频文件")
    if has(r"超过 2 分钟|>2|2 分钟 1 秒", title):
        items.append("已准备时长大于 2 分钟的音频")
    if has(r"不超过 2 分钟|<=2|前 2 分钟|2 分钟整点", title):
        items.append("已准备时长不超过 2 分钟的音频")
    if module == "我的" or has(r"播放页|歌曲|专辑|作品", title):
        items.append("账号下已存在可播放作品")
    if has(r"空状态|暂无可添加", title):
        items.append("账号处于对应空数据状态")
    if has(r"下载", title):
        items.append("当前作品支持下载入口展示")
    if has(r"后台配置|配置", title):
        items.append("后台已下发本场景所需配置")
    return "\n".join(unique(items))


def action_steps(module: str, section: str, title: str) -> str:
    steps = ["打开应用并登录满足前置条件的账号"]
    if module == "首页":
        if section == "入口与默认落点":
            steps.append("冷启动应用，并按标题场景执行首次进入、后台恢复或重新打开操作")
            steps.append("观察首屏落点、旧入口回归表现和纯音乐默认状态")
        elif section == "首页界面翻新":
            steps.append("进入首页新版界面，滚动查看页面入口、卡片、按钮和图文区域")
            steps.append("按标题场景点击关键入口或切换屏幕尺寸进行检查")
        elif has(r"订阅页|价格|付费", title):
            steps.append("从首页触发需要订阅或付费的入口")
            steps.append("观察订阅页、价格展示和返回链路")
        elif section == "有钻石无会员能力边界":
            steps.append("使用非会员账号进入首页相关付费功能入口")
            steps.append("按标题场景分别校验有钻石、无钻石和下载格式限制")
        elif section == "商业化与扣费":
            steps.append("进入首页分轨链路并上传满足场景要求的音频")
            steps.append("按标题场景触发分轨、订阅或钻石消耗流程")
        elif section == "异常与兼容":
            steps.append("进入首页分轨链路并按标题场景模拟权限、弱网、重复点击或配置异常")
            steps.append("观察页面拦截、提示、任务创建和上下文保留情况")
        else:
            steps.append("在首页点击【分轨】入口")
            steps.append("按当前测试标题执行上传、分流、付费或结果查看操作")
    elif module == "创作页":
        steps.append("进入创作页并切换到对应的灵感创作或高级创作模式")
        if has(r"音频参考|上传|录制|版权|Suno", title):
            steps.append("点击【+音频】并按场景选择本地上传、现场录制或播放页带入音频")
        elif has(r"风格|标签|优化", title):
            steps.append("在风格描述区域输入或选择标签，并触发相关按钮")
        elif has(r"模型|会员|订阅", title):
            steps.append("打开模型选择区域并选择对应模型")
        elif has(r"提交|任务|成功|失败|查看|清空", title):
            steps.append("填写满足提交条件的创作内容并点击生成")
        else:
            steps.append("按当前测试标题填写、切换或调整对应字段")
    else:
        if has(r"播放页", section) or has(r"下载|分享|分轨|创作同款", title):
            steps.append("进入作品播放页，打开对应功能入口")
            steps.append("按当前测试标题执行二创、下载、分轨、分享或删除操作")
        else:
            steps.append("进入【我的】-【专辑】页面")
            steps.append("按当前测试标题执行创建、添加、编辑、删除或播放操作")
    steps.append("观察页面状态、提示文案、列表数据、余额/权益变化和跳转结果")
    return "\n".join(f"{index}. {step}" for index, step in enumerate(steps, 1))


def expected_result(title: str) -> str:
    if has(r"字数限制", title):
        return f"{title}；达到限制值时输入仍可保存，超过限制后继续输入被拦截或不再写入，原有内容不丢失。"
    if has(r"范围 0-10|显示范围为 0-10", title):
        return f"{title}；低于最小值、高于最大值和超过 1 位小数的输入被限制，默认值展示正确。"
    if has(r"成功", title):
        return f"{title}；成功提示、目标页面、列表数据和相关状态刷新正确，不产生重复记录或脏数据。"
    if has(r"失败", title):
        return f"{title}；失败提示文案准确，页面停留和可重试状态正确，已定义的余额或数据回退正确。"
    if has(r"不足|无钻石|余额不足", title):
        return f"{title}；系统按 PRD 展示订阅或购买钻石入口，不直接执行需权益或需余额的操作。"
    if has(r"会员|非会员|钻石", title):
        return f"{title}；账号身份、余额判断、按钮展示、跳转入口和扣减结果均与 PRD 分流规则一致。"
    if has(r"下载", title):
        return f"{title}；下载任务状态、横幅提示、文件或相册跳转、失败退钻表现与 PRD 一致。"
    if has(r"上传|录制|权限", title):
        return f"{title}；上传入口、权限触发、上传结果和后续页面状态正确，无重复上传或上下文丢失。"
    if has(r"跳转|返回|回流|恢复|查看", title):
        return f"{title}；目标页面正确，来源参数和任务结果回流正确，返回后不丢失当前上下文。"
    if has(r"展示|显示|隐藏|默认|文案|入口|页面|列表|横幅|弹窗|气泡|Toast", title):
        return f"{title}；对应页面元素、文案、默认值、显示/隐藏状态与 PRD 要求一致。"
    if has(r"删除|二次确认", title):
        return f"{title}；二次确认、取消保留、确认删除和关联数据保留规则均正确。"
    return f"{title}；页面状态、数据变化、按钮可用性和提示结果均与 PRD 对应规则一致。"


def source_remark(leaf: dict) -> str:
    return (
        f"XMind路径：{path_text(leaf)}\n"
        "来源：音创AI/Melod AI-- V1.8.0.md；需求分析与测试理解简报_V1.8.0.md"
    )


def is_abnormal(title: str) -> bool:
    return has(r"失败|拒绝|不足|异常|弱网|重复|取消|超时|不可用|杀进程|错误|无钻石", title)


def is_recovery(title: str) -> bool:
    return has(r"回流|返回|恢复|重进|后台|跳转|刷新|保留|清空|查看|自动执行|继续", title)


def is_boundary(title: str) -> bool:
    return has(r"边界|2 分钟|2分钟|80|200|4500|0-10|50|20 个|1.5 秒|3 秒|超过|不超过|小数|字数|上限|长", title)


def is_branch(title: str) -> bool:
    return has(r"会员|非会员|钻石|歌曲|纯音乐|翻唱|添加人声|添加纯音乐|MP3|MP4|WAV|现有|新建", title)


def is_intercept(title: str) -> bool:
    return has(r"触发订阅|拦截|不显示|不可|置灰|二次确认|权限|拒绝|限制|关闭|取消", title)


def build_rows(trees: dict) -> dict:
    leaves: list[dict] = []
    for tree in trees.values():
        walk_leaves(tree, [], leaves)

    cases: list[list[str]] = []
    mappings: list[list[str]] = []
    pending: list[list[str]] = []
    case_by_path: dict[str, str] = {}
    pending_by_path: dict[str, str] = {}
    case_seq = 1
    pending_seq = 1

    for leaf in leaves:
        module = module_of(leaf)
        section = section_of(leaf)
        if is_pending_title(leaf["title"]) or section == "待确认事项":
            pending_id = f"PEND-{pending_seq:03d}"
            pending_seq += 1
            pending_by_path[path_text(leaf)] = pending_id
            pending.append([
                pending_id,
                module,
                branch_of(leaf),
                leaf["title"],
                "原始需求未定义、存在冲突或设计未落地",
                "澄清后再转化为正式测试用例，避免污染正式预期结果",
                source_remark(leaf),
            ])
            mappings.append([
                path_text(leaf),
                node_type(leaf["title"]),
                "否",
                "",
                f"原始需求未定义或存在冲突，已登记为 {pending_id}",
            ])
            continue

        case_id = f"TC-V180-{case_seq:03d}"
        case_seq += 1
        case_by_path[path_text(leaf)] = case_id
        cases.append([
            case_id,
            PRODUCT,
            f"{module}-{branch_of(leaf)}",
            VERSION,
            f"{section}（子模块）之{scenario_tag(leaf['title'])}（场景）的验证：{leaf['title']}",
            action_steps(module, section, leaf["title"]),
            preconditions(module, leaf["title"]),
            expected_result(leaf["title"]),
            "",
            source_remark(leaf),
            "",
            "",
        ])
        mappings.append([path_text(leaf), node_type(leaf["title"]), "是", case_id, ""])

    return {
        "leaves": leaves,
        "cases": cases,
        "mappings": mappings,
        "pending": pending,
        "case_by_path": case_by_path,
        "pending_by_path": pending_by_path,
    }


def build_coverage(trees: dict, cases: list[list[str]], case_by_path: dict[str, str], pending_by_path: dict[str, str]) -> list[list[str]]:
    rows: list[list[str]] = []
    case_lookup = {row[0]: row for row in cases}
    for tree in trees.values():
        module_node = tree["children"][0]
        for section in module_node["children"]:
            section_leaves: list[dict] = []
            walk_leaves(section, [tree["title"], module_node["title"]], section_leaves)
            ids = [case_by_path[path_text(leaf)] for leaf in section_leaves if path_text(leaf) in case_by_path]
            pending_ids = [pending_by_path[path_text(leaf)] for leaf in section_leaves if path_text(leaf) in pending_by_path]

            def by(predicate: Callable[[str], bool]) -> list[str]:
                return [
                    case_by_path[path_text(leaf)]
                    for leaf in section_leaves
                    if predicate(leaf["title"]) and path_text(leaf) in case_by_path
                ]

            main = [case_id for case_id in ids if not is_abnormal(case_lookup[case_id][4])]
            abnormal = by(is_abnormal)
            recovery = by(is_recovery)
            boundary = by(is_boundary)
            branch = by(is_branch)
            intercept = by(is_intercept)
            pending_reason = f"；相关待确认：{'、'.join(pending_ids)}" if pending_ids else ""

            def value_or_reason(values: list[str], reason: str) -> str:
                return "、".join(values) if values else reason

            rows.append([
                section["title"],
                f"{module_node['title']}-{section['title']}",
                value_or_reason(main, f"不适用：该核心点在 XMind 中未拆出独立主流程叶子节点{pending_reason}"),
                value_or_reason(abnormal, f"不适用：PRD 未定义该核心点的关键异常流程{pending_reason}"),
                value_or_reason(recovery, f"不适用：该核心点无跨页回流、恢复或重进要求{pending_reason}"),
                value_or_reason(boundary, f"不适用：该核心点无数值、时长、数量或文本边界要求{pending_reason}"),
                value_or_reason(branch, f"不适用：该核心点无会员、资源、模式、格式或入口正反分支{pending_reason}"),
                value_or_reason(intercept, f"不适用：该核心点无明确拦截条件{pending_reason}"),
                "、".join(ids) if ids else "无正式用例",
                f"存在待确认事项：{'、'.join(pending_ids)}" if pending_ids else "无",
            ])
    return rows


def build_stats(trees: dict, cases: list[list[str]], mappings: list[list[str]], pending: list[list[str]]) -> list[list[object]]:
    rows: list[list[object]] = [
        ["所属产品", PRODUCT, PROJECT],
        ["测试版本", VERSION, "基于 V1.8.0 PRD、需求分析简报和 XMind 测试点"],
        ["正式测试用例数", len(cases), "不含待确认事项"],
        ["XMind 叶子节点数", len(mappings), "每个叶子节点均有映射"],
        ["待确认事项数", len(pending), "未写入正式预期"],
    ]
    for module in trees:
        module_cases = sum(1 for row in cases if row[2].startswith(f"{module}-"))
        module_mappings = sum(1 for row in mappings if f"> {module} >" in row[0])
        module_pending = sum(1 for row in pending if row[1] == module)
        rows.append([f"{module} 用例数", module_cases, f"叶子节点 {module_mappings} 个，待确认 {module_pending} 个"])
    return rows


def write_sheet(wb: Workbook, name: str, headers: list[str], rows: list[list[object]], widths: list[int]) -> None:
    ws = wb.create_sheet(name)
    ws.append(headers)
    for row in rows:
        ws.append(row)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(name="Arial", size=10, color="FFFFFF", bold=True)
    body_font = Font(name="Arial", size=10)
    thin = Side(style="thin", color="D9E2EC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.font = body_font
            cell.border = border

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

    for idx, width_px in enumerate(widths, 1):
        # Excel width units are approximate; this keeps Python output visually close to the prior workbook.
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = max(10, min(80, width_px / 7.2))

    ws.row_dimensions[1].height = 25.5
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 72

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def build_workbook(trees: dict, built: dict, coverage: list[list[str]], stats: list[list[object]]) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    write_sheet(
        wb,
        "测试用例",
        ["用例编号", "所属产品", "所属模块/链路", "测试版本", "测试标题", "操作步骤", "前置条件", "预期结果", "实际结果", "附件（备注）", "测试负责人", "时间"],
        built["cases"],
        [100, 130, 220, 80, 380, 420, 260, 460, 100, 360, 100, 100],
    )
    write_sheet(
        wb,
        "核心测试点覆盖矩阵",
        ["核心测试点", "所属模块/链路", "主流程", "关键异常", "关键回流/恢复", "关键边界", "关键正反分支", "关键拦截条件", "对应测试用例编号", "不适用说明"],
        coverage,
        [180, 220, 360, 360, 360, 360, 360, 360, 420, 300],
    )
    write_sheet(
        wb,
        "XMind叶子节点映射",
        ["XMind 原节点路径", "节点类型（展示/交互/规则/状态/异常/结果/全局）", "是否生成独立用例", "对应测试用例编号", "未独立生成原因"],
        built["mappings"],
        [520, 180, 120, 180, 420],
    )
    write_sheet(
        wb,
        "待确认事项清单",
        ["待确认编号", "所属模块", "所属链路", "待确认事项", "原因类型", "处理建议", "来源备注"],
        built["pending"],
        [110, 100, 220, 460, 220, 420, 420],
    )
    write_sheet(
        wb,
        "统计",
        ["项目", "数值", "说明"],
        stats,
        [220, 160, 520],
    )
    return wb


def verify_output(path: Path, expected: dict[str, int]) -> None:
    loaded = load_workbook(path, read_only=True)
    required_sheets = ["测试用例", "核心测试点覆盖矩阵", "XMind叶子节点映射", "待确认事项清单", "统计"]
    missing = [sheet for sheet in required_sheets if sheet not in loaded.sheetnames]
    if missing:
        raise RuntimeError(f"缺少工作表：{missing}")
    actual_cases = loaded["测试用例"].max_row - 1
    actual_coverage = loaded["核心测试点覆盖矩阵"].max_row - 1
    actual_mappings = loaded["XMind叶子节点映射"].max_row - 1
    actual_pending = loaded["待确认事项清单"].max_row - 1
    actual = {
        "cases": actual_cases,
        "coverage": actual_coverage,
        "mappings": actual_mappings,
        "pending": actual_pending,
    }
    if actual != expected:
        raise RuntimeError(f"校验失败：expected={expected}, actual={actual}")


def main() -> None:
    trees = testpoint_source.TREES
    built = build_rows(trees)
    coverage = build_coverage(trees, built["cases"], built["case_by_path"], built["pending_by_path"])
    stats = build_stats(trees, built["cases"], built["mappings"], built["pending"])
    wb = build_workbook(trees, built, coverage, stats)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_FILE)
    expected = {
        "cases": len(built["cases"]),
        "coverage": len(coverage),
        "mappings": len(built["mappings"]),
        "pending": len(built["pending"]),
    }
    verify_output(OUTPUT_FILE, expected)
    print({
        "output": str(OUTPUT_FILE),
        **expected,
    })


if __name__ == "__main__":
    main()
